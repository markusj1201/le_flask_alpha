import sys
sys.path.append('../')

def ImportAriesByScenario(scenarioName, start_date, end_date, Area, CorpID = ['ALL']):
    from Model import BPXDatabase as bpxdb
    from Model import QueryFile as qf
    import pandas as pd

    #Create ODS and EDW objects
    Success = True
    Messages = []
    combined_df = pd.DataFrame()
    try:
        ODSobj = bpxdb.GetDBEnvironment('ProdODS', 'OVERRIDE')
        EDWobj = bpxdb.GetDBEnvironment('ProdEDW', 'OVERRIDE')

        scenario_query = qf.ScenarioQuery(scenarioName, CorpID, start_date, end_date)
        results = ODSobj.Query(scenario_query)

        #Extract APINumbers from the results and concatenate into an 'IN' sql clause
        corpid_list = []
        for result in results[0]:
            if not result['CorpID'] in corpid_list:
                corpid_list.append(result['CorpID'])

        if None in corpid_list:
            corpid_list.remove(None)

        key_query = qf.EDWKeyQueryFromCorpID(corpid_list, Area)
        key_results = EDWobj.Query(key_query)

        #Join the key_results to the Aries results
        combined_df = pd.merge(results[1], key_results[1], on='CorpID', how='right')
    
    except Exception as ex:
        Messages.append('Error on Import from Aries. ' + str(ex))
        Success = False

    return combined_df, Success, Messages

def ImportActuals(corpID_list, start_date, end_date, LEName = ''):
    from Model import BPXDatabase as bpxdb
    from Model import QueryFile as qf
    from Model import ModelLayer as m
    from datetime import datetime
    import pandas as pd

    Success = True
    Messages = []
    Actuals = []
    try:
        if isinstance(start_date, datetime):
                start_date = '\'' + start_date.strftime('%Y-%m-%d %H:%M:%S') + '\''

        if isinstance(end_date, datetime):
            end_date = '\'' + end_date.strftime('%Y-%m-%d %H:%M:%S') + '\''
                
        EDWobj = bpxdb.GetDBEnvironment('ProdEDW', 'OVERRIDE')
        query = qf.GetActualsFromDB(corpID_list, start_date, end_date)
        results = EDWobj.Query(query)

        Actuals = results[1]

        #ToDo: Add optional parameter of LE Name. Scan the production adjustments table for any overrides of 
        #potentially incorrect production values from the EDH/EDW tables
        if LEName:
            ProdAdjObj = m.ProductionAdjustments('', [LEName], [], [])        
            ProdAdjsRows, Success, Message = ProdAdjObj.ReadTable()
            if not Success:
                Messages.append(Message)

            #Query the production adjustments and see if there is any well entries for the given LE
            if len(ProdAdjsRows) > 0:
                #Loop through the results of the above query. 
                for row in ProdAdjsRows:  
                #Query the Actuals dataframe for the well and the dates and then update the production value (as long as value is not null)
                    date = row.Date_Key.date()
                    ActualsRow = Actuals.query('CorpID == @row.CorpID and Date_Key == @date')
                    if not ActualsRow.empty:
                        idx = ActualsRow.index.values[0]
                        if row.AdjustedGasProduction:                        
                            Actuals.loc[idx, 'Gas'] = row.AdjustedGasProduction
                        if row.AdjustedOilProduction:
                            Actuals.loc[idx, 'Oil'] = row.AdjustedOilProduction
                        if row.AdjustedWaterProduction:                        
                            Actuals.loc[idx, 'Water'] = row.AdjustedWaterProduction

    except Exception as ex:
        Messages.append('Error during query for actuals data. ' + str(ex))
        Success = False

    return Actuals, Success, Messages

def ImportForecastFromExcel(file, sheetname, IDstart_row, corpId_col, wellName_col, date_row, date_startcol, date_endcol, Phase,  OilNF, GasNF, IDs = ['ALL']):
    import openpyxl as xl
    import pandas as pd

    Success = True
    Messages = []
    return_df = pd.DataFrame()
    try:
        workbook = xl.load_workbook(file, data_only=True)
        worksheet = workbook[sheetname]

        if corpId_col:
            id_col = corpId_col
        else:
            id_col = wellName_col
        
        #Get integer value of parsed Range
        max_row_count = worksheet.max_row
        for row in worksheet.iter_rows(min_row = IDstart_row):
            if corpId_col:
                CorpID = row[id_col -1].value
                WellName = ''
            else:
                WellName = row[id_col -1].value
                CorpID = ''
            wedge = row[id_col].value
            for col in worksheet.iter_cols(min_col = date_startcol, max_col = date_endcol):
                #Add a row to dataframe
                if Phase == 'Gas':
                    Oil = 0
                    Water = 0
                    Gas = worksheet[col[0].column_letter + str(row[0].row)].value
                else:
                    Oil = worksheet[col[0].column_letter + str(row[0].row)].value
                    Water = 0
                    Gas = 0
                Date = worksheet[col[0].column_letter + str(date_row)].value
                return_df = return_df.append({'CorpID': CorpID , 'WellName':WellName, 'Wedge':wedge, 'Date': Date, 'Gas': Gas, 'Oil':Oil, 'Water':Water, 'OilNF' : OilNF, 'GasNF' :GasNF}, ignore_index = True)

    except Exception as ex:
        Messages.append('Error during read from specified LE Spreadsheet. ' + str(ex))
        Success = False

    return return_df, Success, Messages

def ImportGFOFromDB2019(start_date, end_date, WellName_FieldName = ['ALL']):
    from Model import BPXDatabase as bpxdb
    from Model import QueryFile as qf
    import pandas as pd
    from datetime import datetime

    return_df = pd.DataFrame()
    Success = True
    Messages = []
    try:
        if isinstance(start_date, datetime):
                start_date = '\'' + start_date.strftime('%Y-%m-%d %H:%M:%S') + '\''

        if isinstance(end_date, datetime):
            end_date = '\'' + end_date.strftime('%Y-%m-%d %H:%M:%S') + '\''

        TeamOpsObj = bpxdb.GetDBEnvironment('OnPrem', 'OVERRIDE')
        query = qf.GetGFOFromEastDB2019(WellName_FieldName, start_date, end_date)
        results = TeamOpsObj.Query(query)
        return_df = results[1]
    except Exception as ex:
        Messages.append('Error retrieving the GFO data from the desired table. ' + str(ex))
        Success = False

    return return_df, Success, Messages

def ImportGFOFromDB2018(start_date, end_date,  WellFlac = ['ALL']):
    from Model import BPXDatabase as bpxdb
    from Model import QueryFile as qf
    import pandas as pd
    from datetime import datetime

    return_df = pd.DataFrame()
    Success = True
    Messages = []
    try:
        if isinstance(start_date, datetime):
                start_date = '\'' + start_date.strftime('%Y-%m-%d %H:%M:%S') + '\''

        if isinstance(end_date, datetime):
            end_date = '\'' + end_date.strftime('%Y-%m-%d %H:%M:%S') + '\''

        TeamOpsObj = bpxdb.GetDBEnvironment('OnPrem', 'OVERRIDE')
        query = qf.GetGFOFromEastDB2018(WellFlac, start_date, end_date)
        results = TeamOpsObj.Query(query)
        return_df = results[1]
    except Exception as ex:
        Messages.append('Error retrieving the GFO data from the desired table. ' + str(ex))
        Success = False

    return return_df, Success, Messages

def GetWellandCorpID(WellName, CorpID):
    from Model import QueryFile as qf
    from Model import BPXDatabase as bpx

    #Check CorpID if Wellname is passed
    if not CorpID and WellName:
        CorpID_query = qf.EDWKeyQueryFromWellName([WellName])
        EDWObj = bpx.GetDBEnvironment('ProdEDW', 'OVERRIDE')
        res, res_df = EDWObj.Query(CorpID_query)

        if not res_df.empty:
            CorpID = res_df['CorpID'].values[0]

    #Check WellName if CorpID not passed
    if not WellName and CorpID:
        WellName_Query = qf.EDWKeyQueryFromCorpID([CorpID], '')
        EDWObj = bpx.GetDBEnvironment('ProdEDW', 'OVERRIDE')
        res, res_df = EDWObj.Query(WellName_Query)

        if not res_df.empty:
            WellName = res_df['WellName'].values[0]   

    return WellName, CorpID

def GetWedgeData(CorpID, SuppressMessages):
    from Model import QueryFile as qf
    from Model import BPXDatabase as bpx

    from datetime import datetime, timedelta
    import pandas as pd

    Messages = []

    #Get Wedge from First Production Date
    #If an area is passed in as an aggregate, the first production date will be the oldest first production date of its associated well list.
    well_list = GetFullWellList([CorpID])
    first_sales_date_query = qf.FirstProductionDateQuery(well_list)
    first_results = bpx.GetDBEnvironment('ProdEDW', 'OVERRIDE').Query(first_sales_date_query)

    msg = 'Skipped input due to lack of first production date.' + CorpID
    Wedge = ''
    if not first_results[1].empty:
        #check current year and determine if the year of the first production is last year, this year, or base (anything prior to last year)
        prod_date = first_results[1]['FirstProductionDate'].values[0]
        prod_date = pd.to_datetime(prod_date)
        
        if prod_date:
            prod_year = prod_date.year
            this_year = datetime.now().year
            last_year = (datetime.now() - timedelta(days = 365)).year

            if prod_year == this_year:
                Wedge = str(this_year) + ' NWD'
            elif prod_year == last_year:
                Wedge = str(last_year) + ' NWD'
            else:
                Wedge = 'Base'
        else:
            if not SuppressMessages:
                print(msg)
            Messages.append(msg)
            
    else:                        
        Messages.append(msg)
        if not SuppressMessages:
            print(msg)
        
    return Wedge, Messages

def GetFullWellList(well_list):
    from Model import ModelLayer as m
    import pandas as pd

    #Check each item to see if an entry exists as an Area table and return a complete list
    new_list = []
    for well in well_list:
        AreaObj = m.AreaAggregation('', [well], [], [])
        Rows, Success, Message = AreaObj.ReadTable()
        if len(Rows) > 0:
            Rows = pd.DataFrame([vars(s) for s in Rows])
            new_wells = Rows['CorpID'].unique()
            if len(list(new_wells)) == 0:
                print('No wells in Area: ' + well)
            new_list.extend(list(new_wells))
        else:
            new_list.append(well)
    
    return new_list



    


    

